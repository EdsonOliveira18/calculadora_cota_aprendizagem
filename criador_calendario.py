{
  "nbformat": 4,
  "nbformat_minor": 0,
  "metadata": {
    "colab": {
      "provenance": [],
      "authorship_tag": "ABX9TyNgiBIwNFYq3wcGYH6k8vv/",
      "include_colab_link": true
    },
    "kernelspec": {
      "name": "python3",
      "display_name": "Python 3"
    },
    "language_info": {
      "name": "python"
    }
  },
  "cells": [
    {
      "cell_type": "markdown",
      "metadata": {
        "id": "view-in-github",
        "colab_type": "text"
      },
      "source": [
        "<a href=\"https://colab.research.google.com/github/EdsonOliveira18/calculadora_cota_aprendizagem/blob/main/criador_calendario.py\" target=\"_parent\"><img src=\"https://colab.research.google.com/assets/colab-badge.svg\" alt=\"Open In Colab\"/></a>"
      ]
    },
    {
      "cell_type": "code",
      "execution_count": 2,
      "metadata": {
        "id": "U9b7Fqf6WAtw"
      },
      "outputs": [],
      "source": [
        "import pandas as pd\n",
        "import numpy as np\n",
        "from datetime import datetime, timedelta\n",
        "\n",
        "# Dados de entrada\n",
        "inicio = datetime(2023, 1, 1)\n",
        "fim = datetime(2023, 12, 31)\n",
        "feriados = ['2023-01-01', '2023-12-25']  # Adicione mais feriados conforme necessário\n",
        "feriados = [datetime.strptime(date, '%Y-%m-%d') for date in feriados]\n",
        "\n",
        "# Parâmetros\n",
        "dias_empresa = 4\n",
        "dias_capacitacao = 1\n",
        "\n",
        "# Criar DataFrame de datas\n",
        "datas = pd.date_range(start=inicio, end=fim)\n",
        "calendario = pd.DataFrame(datas, columns=['Data'])\n",
        "calendario['DiaSemana'] = calendario['Data'].dt.dayofweek\n",
        "calendario['Atividade'] = 'Empresa'  # Default\n",
        "\n",
        "# Marcar finais de semana\n",
        "calendario.loc[calendario['DiaSemana'] >= 5, 'Atividade'] = 'Fim de Semana'\n",
        "\n",
        "# Marcar feriados\n",
        "calendario.loc[calendario['Data'].isin(feriados), 'Atividade'] = 'Feriado'\n",
        "\n",
        "# Marcar dias de capacitação teórica\n",
        "capacitacao_count = 0\n",
        "for i in range(len(calendario)):\n",
        "    if calendario.loc[i, 'Atividade'] == 'Empresa':\n",
        "        capacitacao_count += 1\n",
        "        if capacitacao_count % 5 == 0:\n",
        "            calendario.loc[i, 'Atividade'] = 'Capacitacao'\n",
        "            capacitacao_count = 0\n",
        "\n",
        "# Exportar para CSV\n",
        "calendario.to_csv('calendario_trabalho.csv', index=False)\n",
        "\n",
        "# Exportar para XLSX com cores\n",
        "import openpyxl\n",
        "from openpyxl.styles import PatternFill\n",
        "\n",
        "# Carregar o DataFrame para o Excel\n",
        "calendario.to_excel('calendario_trabalho.xlsx', index=False)\n",
        "wb = openpyxl.load_workbook('calendario_trabalho.xlsx')\n",
        "ws = wb.active\n",
        "\n",
        "# Definir cores\n",
        "fill_empresa = PatternFill(start_color=\"FFA500\", end_color=\"FFA500\", fill_type=\"solid\")\n",
        "fill_capacitacao = PatternFill(start_color=\"00FF00\", end_color=\"00FF00\", fill_type=\"solid\")\n",
        "fill_feriado = PatternFill(start_color=\"808080\", end_color=\"808080\", fill_type=\"solid\")\n",
        "fill_fim_semana = PatternFill(start_color=\"C0C0C0\", end_color=\"C0C0C0\", fill_type=\"solid\")\n",
        "\n",
        "# Aplicar cores\n",
        "for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):\n",
        "    atividade = row[2].value\n",
        "    if atividade == 'Empresa':\n",
        "        fill = fill_empresa\n",
        "    elif atividade == 'Capacitacao':\n",
        "        fill = fill_capacitacao\n",
        "    elif atividade == 'Feriado':\n",
        "        fill = fill_feriado\n",
        "    else:\n",
        "        fill = fill_fim_semana\n",
        "    for cell in row:\n",
        "        cell.fill = fill\n",
        "\n",
        "wb.save('calendario_trabalho.xlsx')"
      ]
    }
  ]
}